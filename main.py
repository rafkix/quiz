import asyncio
import copy
import logging
import os
import random
from pathlib import Path

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
    PollAnswer,
)
from dotenv import load_dotenv
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# =========================================
# CONFIG
# =========================================

load_dotenv()

TOKEN = os.environ.get("BOT_TOKEN")
if not TOKEN:
    raise RuntimeError(
        "BOT_TOKEN topilmadi. .env faylida BOT_TOKEN=your_token_here deb yozing."
    )

QUIZ_FILE = os.environ.get("QUIZ_FILE", "jismoniy_tarbiya_testlari.xlsx")
FLASHCARD_FILE = os.environ.get("FLASHCARD_FILE", "Data_Analytics_Interview_100.xlsx")

TELEGRAM_MSG_LIMIT = 4000  # 4096 haqiqiy limit, ehtiyot uchun kichikroq

bot = Bot(TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()

# =========================================
# LOAD: MCQ QUIZ (A/B/C/D + poll)
# =========================================


def load_quiz_questions(path: str) -> list[dict]:
    if not Path(path).exists():
        logger.warning("Quiz fayli topilmadi: %s (bu rejim o'chirilgan bo'ladi)", path)
        return []

    wb = load_workbook(path)
    ws = wb.active

    questions = []
    skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            _, question, a, b, c, d, correct, *_ = row

            if not question:
                skipped += 1
                continue

            options = [
                ("A", str(a).strip()),
                ("B", str(b).strip()),
                ("C", str(c).strip()),
                ("D", str(d).strip()),
            ]

            correct = str(correct).strip().upper()

            if correct not in ("A", "B", "C", "D"):
                logger.warning("Quiz row %d: noto'g'ri correct qiymati, o'tkazib yuborildi", row_num)
                skipped += 1
                continue

            random.shuffle(options)
            poll_options = [x[1] for x in options]
            correct_index = next(i for i, x in enumerate(options) if x[0] == correct)

            questions.append(
                {
                    "question": str(question).strip(),
                    "options": poll_options,
                    "correct_index": correct_index,
                }
            )

        except Exception:
            logger.exception("Quiz row %d: qatorni o'qishda xatolik", row_num)
            skipped += 1

    logger.info("Quiz yuklandi: %d ta savol, o'tkazib yuborildi: %d", len(questions), skipped)
    return questions


# =========================================
# LOAD: FLASHCARDS (interview Q&A, ochiq javob)
# Ustunlar: # | Mavzu | Savol | Ideal javob | Nima uchun to'g'ri
#           | Umumiy xatolar | Qo'shimcha savollar | Qiyinlik
# =========================================


def load_flashcards(path: str) -> list[dict]:
    if not Path(path).exists():
        logger.warning("Flashcard fayli topilmadi: %s (bu rejim o'chirilgan bo'ladi)", path)
        return []

    wb = load_workbook(path)
    ws = wb.active

    cards = []
    skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            _, topic, question, ideal_answer, why, mistakes, follow_up, difficulty = row

            if not question:
                skipped += 1
                continue

            cards.append(
                {
                    "topic": str(topic or "").strip(),
                    "question": str(question).strip(),
                    "ideal_answer": str(ideal_answer or "").strip(),
                    "why": str(why or "").strip(),
                    "mistakes": str(mistakes or "").strip(),
                    "follow_up": str(follow_up or "").strip(),
                    "difficulty": str(difficulty or "").strip(),
                }
            )

        except Exception:
            logger.exception("Flashcard row %d: qatorni o'qishda xatolik", row_num)
            skipped += 1

    logger.info("Flashcard yuklandi: %d ta savol, o'tkazib yuborildi: %d", len(cards), skipped)
    return cards


QUIZ_QUESTIONS = load_quiz_questions(QUIZ_FILE)
FLASHCARDS = load_flashcards(FLASHCARD_FILE)

if not QUIZ_QUESTIONS and not FLASHCARDS:
    raise RuntimeError(
        "Hech qanday savol yuklanmadi (quiz ham, flashcard ham). "
        "QUIZ_FILE / FLASHCARD_FILE yo'llarini .env da tekshiring."
    )

# =========================================
# USER SESSIONS
# =========================================

users: dict[int, dict] = {}

# =========================================
# /start -> REJIM TANLASH
# =========================================


@dp.message(Command("start"))
async def start(message: Message):
    buttons = []

    if QUIZ_QUESTIONS:
        buttons.append([InlineKeyboardButton(text="📝 Test (variantli)", callback_data="mode:quiz")])
    if FLASHCARDS:
        buttons.append([InlineKeyboardButton(text="🧠 Flashcard (interview savollari)", callback_data="mode:flashcard")])

    if not buttons:
        await message.answer("⚠️ Hozircha bironta ham savol fayli yuklanmagan.")
        return

    await message.answer(
        "Qaysi rejimda mashq qilamiz?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


# =========================================
# REJIM TANLANDI
# =========================================


@dp.callback_query(F.data == "mode:quiz")
async def choose_quiz(callback: CallbackQuery):
    user_id = callback.from_user.id

    questions = copy.deepcopy(QUIZ_QUESTIONS)
    random.shuffle(questions)

    users[user_id] = {
        "mode": "quiz",
        "questions": questions,
        "index": 0,
        "score": 0,
    }

    await callback.message.edit_text(f"✅ Test boshlandi!\n\nJami savollar: {len(questions)}")
    await callback.answer()

    await send_quiz_question(user_id=user_id, chat_id=callback.message.chat.id)


@dp.callback_query(F.data == "mode:flashcard")
async def choose_flashcard(callback: CallbackQuery):
    user_id = callback.from_user.id

    cards = copy.deepcopy(FLASHCARDS)
    random.shuffle(cards)

    users[user_id] = {
        "mode": "flashcard",
        "questions": cards,
        "index": 0,
    }

    await callback.message.edit_text(f"🧠 Flashcard rejimi boshlandi!\n\nJami savollar: {len(cards)}")
    await callback.answer()

    await send_flashcard(user_id=user_id, chat_id=callback.message.chat.id)


# =========================================
# QUIZ: SAVOL YUBORISH (poll)
# =========================================


async def send_quiz_question(user_id: int, chat_id: int):
    user = users.get(user_id)
    if user is None or user["mode"] != "quiz":
        return

    index = user["index"]

    if index >= len(user["questions"]):
        score = user["score"]
        total = len(user["questions"])

        await bot.send_message(
            chat_id,
            f"🎉 Test tugadi!\n\n"
            f"✅ To'g'ri javoblar: {score}\n"
            f"❌ Noto'g'ri javoblar: {total - score}\n"
            f"📊 Jami: {total}",
        )
        del users[user_id]
        return

    q = user["questions"][index]

    try:
        poll = await bot.send_poll(
            chat_id=chat_id,
            question=q["question"],
            options=q["options"],
            type="quiz",
            correct_option_id=q["correct_index"],
            is_anonymous=False,
            explanation=f"✅ To'g'ri javob: {q['options'][q['correct_index']]}",
        )
    except Exception:
        logger.exception("Poll yuborishda xatolik (user_id=%s, index=%s)", user_id, index)
        await bot.send_message(chat_id, "⚠️ Savolni yuborishda xatolik, keyingisiga o'tamiz.")
        user["index"] += 1
        await send_quiz_question(user_id=user_id, chat_id=chat_id)
        return

    q["poll_id"] = poll.poll.id


@dp.poll_answer()
async def handle_poll_answer(poll_answer: PollAnswer):
    user_id = poll_answer.user.id
    user = users.get(user_id)

    if user is None or user["mode"] != "quiz":
        return

    index = user["index"]
    if index >= len(user["questions"]):
        return

    q = user["questions"][index]

    if q.get("poll_id") != poll_answer.poll_id:
        return

    if not poll_answer.option_ids:
        return

    selected = poll_answer.option_ids[0]

    if selected == q["correct_index"]:
        user["score"] += 1

    user["index"] += 1

    await asyncio.sleep(0.5)
    await send_quiz_question(user_id=user_id, chat_id=user_id)


# =========================================
# FLASHCARD: SAVOL YUBORISH / JAVOB OCHISH
# =========================================


def format_question(card: dict) -> str:
    parts = []
    if card["topic"]:
        parts.append(f"🏷 <b>{card['topic']}</b>" + (f" · {card['difficulty']}" if card["difficulty"] else ""))
    parts.append(f"\n❓ {card['question']}")
    return "\n".join(parts)


def format_answer(card: dict) -> str:
    parts = [format_question(card), ""]

    if card["ideal_answer"]:
        parts.append(f"✅ <b>Ideal javob:</b>\n{card['ideal_answer']}")
    if card["why"]:
        parts.append(f"\n💡 <b>Nega to'g'ri:</b>\n{card['why']}")
    if card["mistakes"]:
        parts.append(f"\n⚠️ <b>Ko'p uchraydigan xatolar:</b>\n{card['mistakes']}")
    if card["follow_up"]:
        parts.append(f"\n🔁 <b>Qo'shimcha savollar:</b>\n{card['follow_up']}")

    return "\n".join(parts)


async def send_flashcard(user_id: int, chat_id: int):
    user = users.get(user_id)
    if user is None or user["mode"] != "flashcard":
        return

    index = user["index"]

    if index >= len(user["questions"]):
        await bot.send_message(chat_id, f"🎉 Barcha savollar tugadi!\n\n📊 Jami: {len(user['questions'])} ta savol ko'rib chiqildi.")
        del users[user_id]
        return

    card = user["questions"][index]

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="👁 Javobni ko'rsatish", callback_data="reveal")]]
    )

    await bot.send_message(chat_id, format_question(card), reply_markup=keyboard)


async def send_long(chat_id: int, text: str):
    """Telegramning 4096 belgi limitidan oshsa, matnni bo'laklarga bo'lib yuboradi."""
    if len(text) <= TELEGRAM_MSG_LIMIT:
        await bot.send_message(chat_id, text)
        return

    for i in range(0, len(text), TELEGRAM_MSG_LIMIT):
        await bot.send_message(chat_id, text[i : i + TELEGRAM_MSG_LIMIT])


@dp.callback_query(F.data == "reveal")
async def reveal_answer(callback: CallbackQuery):
    user_id = callback.from_user.id
    user = users.get(user_id)

    if user is None or user["mode"] != "flashcard":
        await callback.answer()
        return

    index = user["index"]
    if index >= len(user["questions"]):
        await callback.answer()
        return

    card = user["questions"][index]
    answer_text = format_answer(card)

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="➡️ Keyingi savol", callback_data="next")]]
    )

    try:
        if len(answer_text) <= TELEGRAM_MSG_LIMIT:
            await callback.message.edit_text(answer_text, reply_markup=keyboard)
        else:
            # Juda uzun javob - alohida xabar sifatida yuboramiz, tugma ostiga qo'yamiz
            await callback.message.edit_reply_markup(reply_markup=None)
            await send_long(callback.message.chat.id, answer_text)
            await bot.send_message(callback.message.chat.id, "⬆️ Javob yuqorida", reply_markup=keyboard)
    except Exception:
        logger.exception("Javobni ko'rsatishda xatolik (user_id=%s)", user_id)

    await callback.answer()


@dp.callback_query(F.data == "next")
async def next_flashcard(callback: CallbackQuery):
    user_id = callback.from_user.id
    user = users.get(user_id)

    if user is None or user["mode"] != "flashcard":
        await callback.answer()
        return

    user["index"] += 1
    await callback.answer()

    await send_flashcard(user_id=user_id, chat_id=callback.message.chat.id)


# =========================================
# MAIN
# =========================================


async def main():
    logger.info(
        "Bot ishga tushdi... (quiz: %d ta, flashcard: %d ta)",
        len(QUIZ_QUESTIONS),
        len(FLASHCARDS),
    )
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
